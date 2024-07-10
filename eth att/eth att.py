from web3 import Web3
from eth_account import Account
from openpyxl import Workbook, load_workbook
import datetime
import os

# 连接到以太坊节点
w3 = Web3(Web3.HTTPProvider('https://mainnet.infura.io/v3/e658b23b88e34ffead308d481ea1b1cb'))

# 获取程序根目录
root_dir = os.path.dirname(os.path.abspath(__file__))

# Excel 文件路径
eth_data_path = os.path.join(root_dir, 'eth_data.xlsx')
positive_balance_data_path = os.path.join(root_dir, 'positive_balance_data.xlsx')

# 加载现有的Excel文件或创建新文件
try:
    workbook = load_workbook(eth_data_path)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Private Key", "Address", "ETH Balance"])

# 创建另一个Excel文件来保存余额大于0的账户信息
positive_balance_workbook = Workbook()
positive_balance_sheet = positive_balance_workbook.active
positive_balance_sheet.append(["Private Key", "Address", "ETH Balance"])

# 循环十次
for i in range(80):
    for _ in range(1000):
        account = Account.create()

        # 打印生成的密钥信息
        private_key = account._private_key.hex()
        address = account._address

        # 获取账户余额
        balance = w3.eth.get_balance(address)
        eth_balance = w3.from_wei(balance, 'ether')

        print("Private Key:", private_key)
        print("Address:", address)
        print("ETH Balance:", eth_balance)
        print("-----------------------")

        # 将数据添加到Excel文件
        sheet.append([private_key, address, eth_balance])

        # 如果账户余额大于0，则将数据保存到另一个Excel文件
        if eth_balance > 0:
            positive_balance_sheet.append([private_key, address, eth_balance])

    # 保存Excel文件
    workbook.save(eth_data_path)
    positive_balance_workbook.save(positive_balance_data_path)

print("DONE")

# 获取当前时间
current_time = datetime.datetime.now()

# 格式化当前时间
formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
print(formatted_time)
